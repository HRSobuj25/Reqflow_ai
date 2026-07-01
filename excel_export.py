"""
excel_export.py — ReqFlow AI · Professional Excel Export Engine
==============================================================
Converts ERP requirement documents (BRD, SRS, Use Cases, User Stories,
Database Suggestions, KPIs, Workflow, Reports) to structured Microsoft
Excel spreadsheets (.xlsx) using pandas and openpyxl.
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Public constants
# ---------------------------------------------------------------------------
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Style configs
_COLOR_PURPLE_HEX = "8B5CF6"
_COLOR_TH_BG_HEX = "EEE9FF"
_COLOR_TH_TEXT_HEX = "5B21B6"
_COLOR_LIGHT_GRAY_HEX = "F3F4F6"
_COLOR_BORDER_HEX = "E5E7EB"

_SECTION_META: dict[str, tuple[str, str]] = {
    "brd": ("Business Requirements", "BRD"),
    "srs": ("Software Specifications", "SRS"),
    "use_cases": ("Use Case Specifications", "UC"),
    "user_stories": ("User Stories & Acceptance", "US"),
    "db_suggestions": ("Database Design", "DB"),
    "kpis": ("Key Performance Indicators", "KPI"),
    "workflow": ("Workflow Specifications", "WF"),
    "reports": ("Report Specifications", "RPT"),
}

# ===========================================================================
#  Markdown Table Parser
# ===========================================================================
def _parse_markdown_tables(text: str) -> list[tuple[int, pd.DataFrame]]:
    """
    Finds and parses all markdown tables in the text.
    Returns a list of tuples: (line_index_where_table_starts, pandas.DataFrame).
    """
    lines = (text or "").splitlines()
    tables = []
    current_table_lines = []
    start_idx = -1

    for idx, line in enumerate(lines):
        clean_line = line.strip()
        if clean_line.startswith("|") and clean_line.endswith("|"):
            if start_idx == -1:
                start_idx = idx
            current_table_lines.append(clean_line)
        else:
            if current_table_lines:
                df = _convert_table_lines_to_df(current_table_lines)
                if df is not None:
                    tables.append((start_idx, df))
                current_table_lines = []
                start_idx = -1
                
    if current_table_lines:
        df = _convert_table_lines_to_df(current_table_lines)
        if df is not None:
            tables.append((start_idx, df))
            
    return tables


def _convert_table_lines_to_df(lines: list[str]) -> Optional[pd.DataFrame]:
    """Helper to convert raw table pipes to a DataFrame."""
    parsed_rows = []
    for line in lines:
        # Skip separators like |---|---|
        if re.match(r"^\|[\s\-:|]+\|$", line):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        parsed_rows.append(cells)
        
    if len(parsed_rows) < 1:
        return None
        
    headers = parsed_rows[0]
    data = parsed_rows[1:]
    
    # Standardize column lengths
    max_cols = len(headers)
    cleaned_data = []
    for row in data:
        if len(row) < max_cols:
            row += [""] * (max_cols - len(row))
        elif len(row) > max_cols:
            row = row[:max_cols]
        cleaned_data.append(row)
        
    return pd.DataFrame(cleaned_data, columns=headers)


# ===========================================================================
#  Styling Helper
# ===========================================================================
def _style_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Applies high-end typography and formatting to the worksheet."""
    ws.views.sheetView[0].showGridLines = True
    
    thin_side = Side(border_style="thin", color=_COLOR_BORDER_HEX)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Formatting styles
    font_normal = Font(name="Cambria", size=11, color="1F2937")
    font_bold = Font(name="Cambria", size=11, bold=True, color="1F2937")
    font_header = Font(name="Cambria", size=11, bold=True, color=_COLOR_TH_TEXT_HEX)
    fill_header = PatternFill(start_color=_COLOR_TH_BG_HEX, end_color=_COLOR_TH_BG_HEX, fill_type="solid")
    
    # Iterate and format cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if not cell.value:
                continue
                
            cell.font = font_normal
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Format custom section banners
            if isinstance(cell.value, str) and (cell.value.startswith("[") or cell.value.startswith("#")):
                cell.font = Font(name="Cambria", size=14, bold=True, color="1E1B4B")
                ws.row_dimensions[cell.row].height = 28
                continue
                
            # If the cell is in a header/table row, color it
            # (Heuristic: first row of tables or metadata cells)
            if cell.row in ws.merged_cells:
                continue

    # Autofit Column Widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't auto-fit super wide if we only have one column (narrative text)
        if ws.max_column == 1:
            ws.column_dimensions[col_letter].width = 90
            continue
            
        for cell in col:
            val_str = str(cell.value or '')
            # Strip markdown bold/italics markers for length calculation
            val_str = re.sub(r"\*\*|`", "", val_str)
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


# ===========================================================================
#  Worksheet Builder
# ===========================================================================
def _write_section_to_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title: str,
    content: str,
    project_name: str,
    industry: str,
    timestamp: str
) -> None:
    """Writes the markdown block content to the sheet, extracting tables to grids."""
    # Write Title Block
    ws.cell(row=1, column=1, value=title.upper())
    ws.cell(row=1, column=1).font = Font(name="Cambria", size=16, bold=True, color="1E1B4B")
    
    ws.cell(row=2, column=1, value=f"Project: {project_name}  |  Industry: {industry}  |  Generated: {timestamp}")
    ws.cell(row=2, column=1).font = Font(name="Cambria", size=10, italic=True, color="4B5563")
    
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10  # Spacer
    
    tables = _parse_markdown_tables(content)
    lines = content.splitlines()
    
    current_row = 5
    thin_side = Side(border_style="thin", color=_COLOR_BORDER_HEX)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    fill_th = PatternFill(start_color=_COLOR_TH_BG_HEX, end_color=_COLOR_TH_BG_HEX, fill_type="solid")
    font_th = Font(name="Cambria", size=11, bold=True, color=_COLOR_TH_TEXT_HEX)
    
    # Track which lines were tables so we skip printing them as plain lines
    table_lines_mask = [False] * len(lines)
    for start_line, df in tables:
        # Mask matching lines in source
        # A markdown table length includes headers, dividers, and data rows
        tbl_len = len(df) + 2  # df rows + header row + separator line
        for offset in range(tbl_len):
            if start_line + offset < len(table_lines_mask):
                table_lines_mask[start_line + offset] = True

    line_idx = 0
    while line_idx < len(lines):
        # Check if table starts here
        table_match = next((t for t in tables if t[0] == line_idx), None)
        
        if table_match:
            _, df = table_match
            # Write Table to sheet
            headers = list(df.columns)
            
            # Write Header Row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = font_th
                cell.fill = fill_th
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[current_row].height = 24
            current_row += 1
            
            # Write Data Rows
            for _, row in df.iterrows():
                for col_idx, val in enumerate(row, 1):
                    # Clean bold formatting from values
                    clean_val = re.sub(r"\*\*|`", "", str(val))
                    cell = ws.cell(row=current_row, column=col_idx, value=clean_val)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
            current_row += 1  # spacer after table
            line_idx += len(df) + 2
        else:
            line = lines[line_idx].strip()
            # If not masked and is not blank
            if not table_lines_mask[line_idx] and line:
                # Basic parsing for lists/headings
                clean_val = re.sub(r"\*\*|`", "", line)
                cell = ws.cell(row=current_row, column=1, value=clean_val)
                
                # Check for markdown headers
                if line.startswith("#"):
                    level = len(line) - len(line.lstrip("#"))
                    cell.value = line.replace("#", "").strip()
                    cell.font = Font(name="Cambria", size=13 if level > 1 else 14, bold=True, color="8B5CF6" if level > 1 else "1E1B4B")
                    ws.row_dimensions[current_row].height = 22
                else:
                    ws.row_dimensions[current_row].height = 18
                    
                current_row += 1
            elif not line:
                # Add spacing for paragraph breaks
                current_row += 1
                
            line_idx += 1
            
    _style_sheet(ws)


# ===========================================================================
#  Public API
# ===========================================================================
def build_excel_document(
    project_name: str,
    industry: str,
    content: str,
    doc_label: str,
    timestamp: Optional[str] = None
) -> bytes:
    """
    Builds a single-section Excel document (.xlsx).
    """
    ts = timestamp or datetime.now().strftime("%B %d, %Y at %H:%M")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doc_label[:30] # Excel tab limit is 31 chars
    
    _write_section_to_sheet(ws, doc_label, content, project_name, industry, ts)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def build_complete_suite_excel(
    project_name: str,
    industry: str,
    timestamp: Optional[str],
    sections: dict[str, str]
) -> bytes:
    """
    Builds a multi-tab workbook compiling all requirements sheets.
    """
    ts = timestamp or datetime.now().strftime("%B %d, %Y at %H:%M")
    wb = openpyxl.Workbook()
    
    # ── TAB 1: Summary Cover Sheet
    ws_cover = wb.active
    ws_cover.title = "Summary Overview"
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover.cell(row=1, column=1, value="ReqFlow AI Enterprise Requirement Suite")
    ws_cover.cell(row=1, column=1).font = Font(name="Cambria", size=16, bold=True, color="1E1B4B")
    
    meta_rows = [
        ("Project Name", project_name),
        ("Industry Domain", industry),
        ("Suite Volume", "Complete Requirement Suite (8 Sections)"),
        ("Generated Timestamp", ts),
        ("System Tool", "ReqFlow AI Enterprise Module"),
    ]
    
    thin_side = Side(border_style="thin", color=_COLOR_BORDER_HEX)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    fill_lbl = PatternFill(start_color=_COLOR_TH_BG_HEX, end_color=_COLOR_TH_BG_HEX, fill_type="solid")
    font_lbl = Font(name="Cambria", size=11, bold=True, color=_COLOR_TH_TEXT_HEX)
    
    curr_r = 3
    for lbl, val in meta_rows:
        c1 = ws_cover.cell(row=curr_r, column=1, value=lbl)
        c2 = ws_cover.cell(row=curr_r, column=2, value=val)
        c1.font = font_lbl
        c1.fill = fill_lbl
        c1.border = thin_border
        c2.border = thin_border
        ws_cover.row_dimensions[curr_r].height = 20
        curr_r += 1
        
    # Table of contents index
    curr_r += 2
    ws_cover.cell(row=curr_r, column=1, value="DOCUMENT SPECIFICATIONS INDEX")
    ws_cover.cell(row=curr_r, column=1).font = Font(name="Cambria", size=12, bold=True, color="8B5CF6")
    curr_r += 1
    
    toc_idx = 1
    for key, val in _SECTION_META.items():
        content = sections.get(key, "")
        if content and content.strip() and not content.startswith("*"):
            title = val[0]
            ws_cover.cell(row=curr_r, column=1, value=f"{toc_idx}. {title}")
            ws_cover.cell(row=curr_r, column=2, value=f"Active Workspace Tab Sheet: [{val[1]}]")
            ws_cover.row_dimensions[curr_r].height = 18
            toc_idx += 1
            curr_r += 1
            
    _style_sheet(ws_cover)
    ws_cover.column_dimensions["A"].width = 32
    ws_cover.column_dimensions["B"].width = 45

    # ── TABS 2-9: Requirements Sections
    for key, val in _SECTION_META.items():
        content = sections.get(key, "")
        if content and content.strip() and not content.startswith("*"):
            title, tab_title = val
            ws = wb.create_sheet(title=tab_title)
            _write_section_to_sheet(ws, title, content, project_name, industry, ts)
            
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
